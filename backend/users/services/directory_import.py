import re
from dataclasses import dataclass

from django.contrib.auth import get_user_model
from django.db import transaction
from openpyxl import load_workbook

from users.models import DirectoryImportBatch, DirectoryImportEntry

User = get_user_model()

HEADER_MAP = {
    'primer nombre': 'first_name',
    'nombre': 'first_name',
    'nombres': 'first_name',
    'primer apellido': 'last_name',
    'apellido': 'last_name',
    'segundo apellido': 'second_lastname',
    'documento': 'document_number',
    'numero de documento': 'document_number',
    'número de documento': 'document_number',
    'cedula': 'document_number',
    'cédula': 'document_number',
    'correo': 'email',
    'correo electronico': 'email',
    'correo electrónico': 'email',
    'email': 'email',
    'correo personal': 'personal_email',
    'personal email': 'personal_email',
    'celular': 'phone_number',
    'telefono': 'phone_number',
    'teléfono': 'phone_number',
}

RESTORE_FIELDS = [
    'username', 'password', 'first_name', 'last_name', 'email', 'role', 'roles',
    'document_number', 'second_name', 'second_lastname', 'personal_email',
    'phone_number', 'is_active', 'is_directory_imported', 'requires_onboarding',
    'directory_batch_id',
]


@dataclass
class ImportResult:
    batch_id: int
    created: int
    updated: int
    skipped: int
    errors: list


def normalize_header(value):
    text = re.sub(r'\s+', ' ', str(value or '').strip().lower())
    return HEADER_MAP.get(text)


def clean_cell(value):
    if value is None:
        return ''
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).strip()


def parse_rows(file_obj):
    workbook = load_workbook(file_obj, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [normalize_header(item) for item in rows[0]]
    parsed = []
    for index, row in enumerate(rows[1:], start=2):
        item = {key: clean_cell(row[pos]) for pos, key in enumerate(headers) if key}
        if any(item.values()):
            item['_row'] = index
            parsed.append(item)
    return parsed


def validate_row(item):
    required = ['document_number']
    missing = [field for field in required if not item.get(field)]
    if missing:
        return f"Fila {item['_row']}: faltan {', '.join(missing)}"
    if item.get('email') and '@' not in item['email']:
        return f"Fila {item['_row']}: correo inválido"
    if not item['document_number'].isdigit():
        return f"Fila {item['_row']}: documento debe ser numérico"
    return None


@transaction.atomic
def import_student_directory(file_obj, created_by=None):
    batch = DirectoryImportBatch.objects.create(
        file_name=getattr(file_obj, 'name', 'directorio.xlsx')[:255],
        created_by=created_by if getattr(created_by, 'is_authenticated', False) else None,
    )
    created = updated = skipped = 0
    errors = []
    for item in parse_rows(file_obj):
        error = validate_row(item)
        if error:
            skipped += 1
            errors.append(error)
            create_entry(batch, item, 'skipped', message=error)
            continue
        result = upsert_student(item, batch)
        if result in ('created', 'updated'):
            created += result == 'created'
            updated += result == 'updated'
        else:
            skipped += 1
            errors.append(result)
            create_entry(batch, item, 'skipped', message=result)
    batch.created_count = created
    batch.updated_count = updated
    batch.skipped_count = skipped
    batch.errors = errors[:50]
    batch.save(update_fields=['created_count', 'updated_count', 'skipped_count', 'errors'])
    return ImportResult(batch.id, created, updated, skipped, errors[:50])


def upsert_student(item, batch):
    email = item.get('email', '').lower()
    document = item['document_number']
    document_owner = User.objects.filter(document_number=document).first()
    user = document_owner or (User.objects.filter(email__iexact=email).first() if email else None)
    error = validate_user_match(item, user, document_owner)
    if error:
        return error
    previous = snapshot_user(user) if user else {}
    if not user:
        user = User(username=f'dir-{document}', email=email)
        user.set_unusable_password()
        state = 'created'
    else:
        state = 'updated'
    assign_directory_data(user, item, batch)
    user.save()
    create_entry(batch, item, state, user=user, previous_data=previous)
    return state


def validate_user_match(item, user, document_owner):
    email = item.get('email', '').lower()
    email_owner = User.objects.filter(email__iexact=email).first() if email else None
    if document_owner and email_owner and document_owner.id != email_owner.id:
        return f"Fila {item['_row']}: documento pertenece a otro correo"
    if user and user.role not in ('STUDENT', 'ESTUDIANTE') and 'STUDENT' not in (user.roles or []):
        return f"Fila {item['_row']}: el correo pertenece a un usuario no estudiante"
    return None


def snapshot_user(user):
    return {field: getattr(user, field) for field in RESTORE_FIELDS if field != 'directory_batch_id'} | {
        'directory_batch_id': user.directory_batch_id,
    }


def assign_directory_data(user, item, batch):
    if item.get('first_name'):
        user.first_name = item['first_name']
    if item.get('last_name'):
        user.last_name = item['last_name']
    if item.get('second_name'):
        user.second_name = item['second_name']
    if item.get('second_lastname'):
        user.second_lastname = item.get('second_lastname', '')
    user.document_number = item['document_number']
    if item.get('email'):
        user.username = item['email'].lower()
        user.email = item['email'].lower()
    elif not user.username:
        user.username = f"dir-{item['document_number']}"
    if item.get('personal_email'):
        user.personal_email = item['personal_email'].lower()
    if item.get('phone_number'):
        user.phone_number = item['phone_number']
    user.role = 'STUDENT'
    user.roles = ['STUDENT']
    user.is_active = True
    user.is_directory_imported = True
    user.requires_onboarding = True
    user.directory_batch = batch


def create_entry(batch, item, action, user=None, previous_data=None, message=''):
    DirectoryImportEntry.objects.create(
        batch=batch,
        user=user,
        action=action,
        email=item.get('email', ''),
        document_number=item.get('document_number', ''),
        previous_data=previous_data or {},
        message=message,
    )


@transaction.atomic
def revert_directory_batch(batch):
    if batch.is_reverted:
        return 0
    affected = 0
    for entry in batch.entries.select_related('user').filter(action__in=['created', 'updated']):
        if entry.action == 'created' and entry.user and entry.user.directory_batch_id == batch.id:
            entry.user.delete()
            affected += 1
        elif entry.action == 'updated' and entry.user:
            restore_user(entry.user, entry.previous_data)
            affected += 1
    batch.is_reverted = True
    batch.save(update_fields=['is_reverted'])
    return affected


def restore_user(user, data):
    for field in RESTORE_FIELDS:
        setattr(user, field, data.get(field))
    user.save()
